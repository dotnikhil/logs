import re
import subprocess
import logging
import json
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import os
from flask import Flask, request, jsonify
import requests

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Flask app
app = Flask(__name__)

# Load configuration from file
CONFIG_FILE = "/home/whitelist/config.json"
try:
    with open(CONFIG_FILE, "r") as f:
        config = json.load(f)
except FileNotFoundError:
    logger.error(f"Config file {CONFIG_FILE} not found")
    raise SystemExit(f"Config file {CONFIG_FILE} not found")
except json.JSONDecodeError:
    logger.error(f"Invalid JSON in {CONFIG_FILE}")
    raise SystemExit(f"Invalid JSON in {CONFIG_FILE}")

# Validate required config keys
required_keys = ["WAF_CONFIG", "EXCEL_FILE", "ERROR_EXCEL_FILE", "COLUMNS", "TEAMS_WEBHOOK_URL"]
missing_keys = [key for key in required_keys if key not in config]
if missing_keys:
    logger.error(f"Missing config keys: {missing_keys}")
    raise SystemExit(f"Missing config keys: {missing_keys}")

WAF_CONFIG = config["WAF_CONFIG"]
EXCEL_FILE = config["EXCEL_FILE"]
ERROR_EXCEL_FILE = config["ERROR_EXCEL_FILE"]
COLUMNS = config["COLUMNS"]
TEAMS_WEBHOOK_URL = config["TEAMS_WEBHOOK_URL"]
SHEET2API_KEY = config.get("SHEET2API_KEY", "")  # Optional for cloud logging

def init_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Whitelist Log" if file_path == EXCEL_FILE else "Whitelist Error Log"
        for col, header in enumerate(COLUMNS, 1):
            ws[f"{get_column_letter(col)}1"] = header
        wb.save(file_path)
    return wb

def log_to_excel(ip, environments, requester, purpose, status="Success"):
    file_path = EXCEL_FILE if status == "Success" else ERROR_EXCEL_FILE
    data = {
        "Prod": "y" if "prod" in environments else "",
        "UAT": "y" if "uat" in environments else "",
        "China": "y" if "china" in environments else "",
        "IP": f"{ip}/32" if ip else ip,
        "Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Requested By": requester,
        "XtraNet": "xtranet",
        "Purpose": purpose
    }
    # Try cloud logging with Sheet2API if configured
    if SHEET2API_KEY and file_path.startswith("https://"):
        headers = {"Authorization": f"Bearer {SHEET2API_KEY}"}
        try:
            response = requests.post(file_path, json=data, headers=headers)
            response.raise_for_status()
            logger.info(f"Logged to {'Success' if status == 'Success' else 'Error'} API: {data}")
            return
        except requests.RequestException as e:
            logger.error(f"Failed to log to API: {str(e)}")
    # Fallback to local Excel logging
    wb = init_excel(EXCEL_FILE if status == "Success" else ERROR_EXCEL_FILE)
    ws = wb.active
    row = ws.max_row + 1
    for col, value in enumerate(data.values(), 1):
        ws[f"{get_column_letter(col)}{row}"] = value
    if status == "Failed":
        ws[f"{get_column_letter(8)}{row}"].comment = openpyxl.comments.Comment("Failed to whitelist", "bot")
    wb.save(EXCEL_FILE if status == "Success" else ERROR_EXCEL_FILE)
    logger.info(f"Logged to {'Success' if status == 'Success' else 'Error'} Excel: {data}")

def send_teams_response(message, status="Success"):
    if not TEAMS_WEBHOOK_URL:
        logger.error("TEAMS_WEBHOOK_URL not set in config file")
        return
    card = {
        "type": "AdaptiveCard",
        "body": [
            {
                "type": "TextBlock",
                "size": "Medium",
                "weight": "Bolder",
                "text": f"IP Whitelisting {'Success' if status == 'Success' else 'Error'}"
            },
            {
                "type": "TextBlock",
                "text": message,
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S IST')}",
                "size": "Small",
                "color": "Default"
            }
        ],
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "version": "1.0"
    }
    payload = {
        "type": "message",
        "attachments": [
            {
                "contentType": "application/vnd.microsoft.card.adaptive",
                "content": card
            }
        ]
    }
    try:
        response = requests.post(TEAMS_WEBHOOK_URL, json=payload)
        response.raise_for_status()
        logger.info("Sent response to Teams")
    except requests.RequestException as e:
        logger.error(f"Failed to send Teams response: {str(e)}")

def whitelist_ip(message, is_teams=False, requester="unknown"):
    # Capture print output for Teams
    import sys
    from io import StringIO
    old_stdout = sys.stdout
    sys.stdout = mystdout = StringIO()

    try:
        # Check if message starts with !wl
        if not message.lower().startswith("!wl"):
            print("Invalid format. Use: !wl <ip-address> -p <uat prod both> -r <purpose>")
            log_to_excel("", [], requester, "Invalid format", status="Failed")
            return

        # Regex to parse message: !wl <ip-address> -p <uat prod both> -r <purpose>
        pattern = r"^!wl\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\s+-p\s+([a-z\s]+)(?:\s+-r\s+(.+))?$"
        match = re.match(pattern, message.lower().strip())
        
        if not match:
            error_msg = "Invalid format. Use: !wl <ip-address> -p <uat prod both> -r <purpose>"
            print(error_msg)
            log_to_excel("", [], requester, "Invalid format", status="Failed")
            return

        ip_address = match.group(1)
        platform_input = match.group(2).split()
        purpose = match.group(3) if match.group(3) else None

        # Check for missing purpose
        if not purpose:
            error_msg = "Please specify a purpose."
            print(error_msg)
            log_to_excel(ip_address, [], requester, "Missing purpose", status="Failed")
            return

        # Map platform aliases
        valid_envs = {"uat", "prod"}  # Removed 'china'
        environments = []
        for platform in platform_input:
            if platform == "both":
                environments.extend(["uat", "prod"])
            elif platform == "all":
                environments.extend(["uat", "prod"])  # Exclude 'china' from 'all'
            elif platform in valid_envs:
                environments.append(platform)

        # Remove duplicates while preserving order
        environments = list(dict.fromkeys(environments))

        # Check for missing or invalid platforms
        if not environments:
            error_msg = "Please specify at least one platform (uat, prod, or both)."
            print(error_msg)
            log_to_excel(ip_address, [], requester, purpose, status="Failed")
            return
        if not all(env in valid_envs for env in environments):
            error_msg = f"Invalid environment(s). Use: uat, prod, or both"
            print(error_msg)
            log_to_excel(ip_address, [], requester, purpose, status="Failed")
            return

        # Validate IP address (ensure octets are 0-255)
        ip_pattern = r"^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$"
        if not re.match(ip_pattern, ip_address):
            error_msg = f"Invalid IP address: {ip_address}"
            print(error_msg)
            log_to_excel(ip_address, [], requester, purpose, status="Failed")
            return

        # Whitelist IP using AWS CLI
        results = []
        for env in environments:
            # Find the account ID for the environment
            account_id = None
            for acc_id, envs in WAF_CONFIG.items():
                if env in envs:
                    account_id = acc_id
                    break
            if not account_id:
                raise Exception(f"No account found for environment: {env}")

            ip_set_name = WAF_CONFIG[account_id][env]["ip_set_name"]
            ip_set_id = WAF_CONFIG[account_id][env]["ip_set_id"]
            region = WAF_CONFIG[account_id][env]["region"]
            scope = WAF_CONFIG[account_id][env]["scope"]
            new_ip = f"{ip_address}/32"

            # Get current IP set
            get_cmd = f"aws wafv2 get-ip-set --name {ip_set_name} --scope {scope} --id {ip_set_id} --region {region}"
            get_result = subprocess.run(get_cmd, shell=True, capture_output=True, text=True)
            if get_result.returncode != 0:
                raise Exception(f"Failed to get IP set for {env}: {get_result.stderr}")

            ip_set_data = json.loads(get_result.stdout)
            current_ips = ip_set_data["IPSet"]["Addresses"]
            lock_token = ip_set_data["LockToken"]

            # Add new IP if not already present
            if new_ip not in current_ips:
                current_ips.append(new_ip)
                ip_list = " ".join(f"'{ip}'" for ip in current_ips)
                update_cmd = (
                    f"aws wafv2 update-ip-set --name {ip_set_name} --scope {scope} --id {ip_set_id} "
                    f"--region {region} --addresses {ip_list} --lock-token {lock_token}"
                )
                update_result = subprocess.run(update_cmd, shell=True, capture_output=True, text=True)
                if update_result.returncode != 0:
                    raise Exception(f"Failed to update IP set for {env}: {update_result.stderr}")
                results.append(f"Whitelisted {new_ip} in {env} IP set")
            else:
                results.append(f"{new_ip} already in {env} IP set")
            
            logger.info(f"Processed {new_ip} for {env}: {purpose}")

        # Log to Excel
        log_to_excel(ip_address, environments, requester, purpose)

        # Print and send results to Teams (commented out for testing)
        response_text = "\n".join(results)
        print(response_text)
        # if is_teams:
        #     send_teams_response(response_text, status="Success")

    except Exception as e:
        logger.error(f"Error whitelisting IP: {str(e)}")
        error_msg = f"Error: {str(e)}. Please try again or contact the DevOps Team."
        print(error_msg)
        log_to_excel(ip_address, environments, requester, purpose, status="Failed")
        # if is_teams:
        #     send_teams_response(error_msg, status="Failed")
    
    finally:
        # Restore stdout and return output for Teams
        sys.stdout = old_stdout
        return mystdout.getvalue()

@app.route("/webhook", methods=["POST"])
def teams_webhook():
    try:
        data = request.get_json()
        if not data or "text" not in data:
            logger.error("Invalid Teams webhook payload")
            return jsonify({"error": "Invalid payload"}), 400
        
        message = data["text"]
        # Extract requester name from Teams payload
        requester = data.get("from", {}).get("user", {}).get("displayName", "unknown")
        logger.info(f"Received Teams webhook message from {requester}: {message}")
        response_text = whitelist_ip(message, is_teams=True, requester=requester)
        return jsonify({"status": "processed", "message": response_text}), 200
    except Exception as e:
        logger.error(f"Error processing Teams webhook: {str(e)}")
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    logger.info("Starting Flask server for Teams webhook on port 5000...")
    app.run(host="0.0.0.0", port=5000)
